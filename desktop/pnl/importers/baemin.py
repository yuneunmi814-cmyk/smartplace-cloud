"""배민 정산명세서 importer — 2026-05 실파일(복호화)로 검증 완료.

확정 사실:
  - 파일은 CDFV2 Encrypted(암호화). msoffcrypto-tool 로 비번 복호화 후 파싱.
  - 시트 2개: '요약'(월 카테고리 합계) / '상세'(입금 건별, 4단 병합헤더, 32열, 데이터 행5~).
  - 검증: 행별 col5~29 합 = 입금금액(col30), 66건 오차 0. 요약 = 상세 컬럼그룹 합.
  - ★부호규칙이 쿠팡과 반대: 배민은 파일에 수수료/할인이 '이미 음수'. → 그대로(as-is) emit.
  - P&L 제외 항목: 부가세(col26, 부채), 만나서결제 현금차감(col23/24, 현금정산 reconcile).
  - 검증 등식: 매출 − 비용(P&L) = 영업이익; 영업이익 − 부가세 − 만나서결제차감 = 입금액.
"""
from __future__ import annotations

import tempfile

import openpyxl

from .base import DeliveryImporter, ImportResult
import pandas as pd

# 0-based 컬럼 인덱스 (복호화 평문 '상세' 시트, 실파일 검증)
COL = dict(
    deposit_date=0, period=1, deposit_amt=2, service=3, order_type=4,
    order_paid=5, order_meet=6, partial_refund=7,          # 매출(주문금액)
    mid_baemin1=8, mid_alttle=9, mid_gage=10, mid_pickup=11,  # 중개이용료
    disc_immediate=12, disc_menu=13,                        # 고객할인비용(판촉)
    tip_paid=14, tip_meet=15,                               # 가게배달팁(운반비 상쇄)
    club_h=16, club_h_sup=17, club_a=18, club_a_sup=19,     # 배민클럽(순0, 판촉)
    deliv_baemin1=20, deliv_alttle=21,                      # 배달비(운반비)
    pg_fee=22,                                              # 결제정산수수료(지급수수료)
    meet_order_deduct=23, meet_tip_deduct=24,               # 만나서결제 현금차감(P&L 제외)
    deliv_support=25,                                       # 배달비환급프로모션(운반비 상쇄)
    vat=26,                                                 # 부가세(P&L 제외)
    wgg_fee=27, wgg_vat=28,                                 # 우리가게클릭(광고선전비)
    baemin_order=29, deposit_h=30, status=31,
)
DATA_START_ROW = 6  # openpyxl 1-based (헤더 행0~4)


def is_encrypted_ole(filepath: str) -> bool:
    with open(filepath, "rb") as f:
        return f.read(8) == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def decrypt_to_xlsx(filepath: str, password: str) -> str:
    import msoffcrypto

    with open(filepath, "rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password=password)
        out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        office.decrypt(out)
        out.flush()
        return out.name


class BaeminPasswordRequired(RuntimeError):
    """배민 파일 비번이 필요/오류일 때 — UI에서 비번 입력 유도용."""


class BaeminImporter(DeliveryImporter):
    platform = "baemin"
    doc_type = "settlement"

    def __init__(self, password: str | None = None):
        self.password = password

    def identify(self, filepath: str) -> bool:
        # 배민 정산명세서는 암호화 OLE 컨테이너 → 이 시그니처로 식별
        return is_encrypted_ole(filepath)

    def extract(self, filepath: str) -> ImportResult:
        if not self.password:
            raise BaeminPasswordRequired(
                "배민 정산명세서는 암호화되어 있습니다. 비밀번호를 입력해 주세요.")
        try:
            plain = decrypt_to_xlsx(filepath, self.password)
        except Exception as e:
            raise BaeminPasswordRequired(f"배민 복호화 실패(비번 확인): {e}")

        wb = openpyxl.load_workbook(plain, read_only=True, data_only=True)
        if "상세" not in wb.sheetnames:
            raise ValueError("배민 '상세' 시트를 찾을 수 없습니다.")
        ws = wb["상세"]

        def num(v):
            return self.won(v)

        recs, payout, period = [], 0.0, ""
        for r in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            if r[COL["deposit_date"]] in (None, ""):
                continue
            status = str(r[COL["status"]] or "")
            if status and "입금완료" not in status:
                continue  # 요약과 동일하게 입금완료만 집계
            ddate = str(r[COL["deposit_date"]])[:10]
            if not period and len(ddate) >= 7:
                period = ddate[:7]
            order_no = ""

            def emit(item, amt, basis="gross", vat=0.0):
                if amt:
                    recs.append(dict(date=ddate, platform=self.platform,
                                     doc_type=self.doc_type, order_no=order_no,
                                     item_name=item, amount=amt,
                                     vat_basis=basis, vat=vat))

            # 배민은 부호 그대로(as-is): 매출+, 수수료/할인은 이미 −
            # 배민 수수료(중개/결제/배달비)는 '공급가액'(부가세는 별도 col26) → supply.
            # 우리가게클릭은 파일에 부가세 명시(col28) → exact. 매출은 총액 → gross.
            emit("배민 주문금액",
                 num(r[COL["order_paid"]]) + num(r[COL["order_meet"]])
                 + num(r[COL["partial_refund"]]), "gross")
            emit("배민 중개이용료",
                 num(r[COL["mid_baemin1"]]) + num(r[COL["mid_alttle"]])
                 + num(r[COL["mid_gage"]]) + num(r[COL["mid_pickup"]]), "supply")
            emit("배민 결제정산수수료", num(r[COL["pg_fee"]]), "supply")
            emit("배민 고객할인",
                 num(r[COL["disc_immediate"]]) + num(r[COL["disc_menu"]]), "none")
            emit("배민 배민클럽할인",  # 순액 0이지만 명시적으로(할인/지원 쌍)
                 num(r[COL["club_h"]]) + num(r[COL["club_h_sup"]])
                 + num(r[COL["club_a"]]) + num(r[COL["club_a_sup"]]), "none")
            emit("배민 배달팁",
                 num(r[COL["tip_paid"]]) + num(r[COL["tip_meet"]]), "none")
            emit("배민 배달비",
                 num(r[COL["deliv_baemin1"]]) + num(r[COL["deliv_alttle"]]), "supply")
            emit("배민 배달비지원", num(r[COL["deliv_support"]]), "none")
            emit("배민 우리가게클릭광고비",
                 num(r[COL["wgg_fee"]]) + num(r[COL["wgg_vat"]]), "exact",
                 vat=num(r[COL["wgg_vat"]]))           # 우리가게클릭 부가세(col28)
            # P&L 제외: 부가세(26), 만나서결제 현금차감(23/24)
            payout += num(r[COL["deposit_amt"]])

        rows = pd.DataFrame(recs)
        return ImportResult(
            platform=self.platform, doc_type=self.doc_type, period=period,
            rows=rows, payout_reported=payout,
            notes="부가세·만나서결제 현금차감은 P&L 제외(부채/현금정산). 영업이익−부가세−현금차감=입금액.")
