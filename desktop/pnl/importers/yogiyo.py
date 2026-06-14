"""요기요 정산 importer — ⚠️ 실파일 미확보. '헤더 이름(+동의어) 매칭' 적응형 구현.

배민/쿠팡은 실파일로 고정 인덱스를 박았지만, 요기요는 실파일이 없어 컬럼 위치를 단언할 수 없다.
대신:
  1) 헤더 행을 자동 탐지(요기요 키워드가 가장 많은 행)
  2) 컬럼을 동의어로 개념(매출/중개/결제/배달/광고/할인/입금)에 매핑
  3) 비용 컬럼의 부호를 합계로 자동 판별(양수저장이면 −로 환산, 취소행 상계 보존)
  4) 자동 검산: 모든 항목 합 ≈ 입금액 → notes에 '검산 통과/불일치' 기록
실파일이 오면 검산 결과만 보면 맞는지 바로 안다. 안 맞으면 SYNONYMS만 손보면 된다.

근거(개념): 요기요 정산금액 = 주문금액 − 사장님할인 − 주문중개이용료 − 배달서비스이용료
            − 외부결제수수료 − 광고상품비 − 기타. (partner.yogiyo.co.kr/guide/pay)
"""
from __future__ import annotations

import openpyxl
import pandas as pd

from .base import DeliveryImporter, ImportResult

# 개념 → 헤더에 포함될 수 있는 문자열(우선순위 순). 위에서부터 먼저 claim.
SYNONYMS: list[tuple[str, str, list[str]]] = [
    # (개념키, 표시명, [헤더 동의어])  — vat_basis 는 아래 BASIS 참조
    ("deposit",    "요기요 입금액",       ["입금액", "지급액", "정산금액", "정산 금액"]),
    ("brokerage",  "요기요 중개이용료",    ["주문중개이용료", "중개이용료", "중개수수료", "중개"]),
    ("payment",    "요기요 결제수수료",    ["외부결제수수료", "결제수수료", "결제 수수료", "PG수수료"]),
    ("delivery",   "요기요 배달서비스이용료", ["배달서비스이용료", "배달대행", "배달비", "배달 서비스"]),
    ("ad",         "요기요 광고비",       ["광고상품", "광고비", "광고", "요타임딜", "구독료"]),
    ("discount",   "요기요 할인",         ["사장님할인", "할인", "쿠폰", "프로모션"]),
    ("vat",        "요기요 부가세",       ["부가세", "부가가치세"]),
    ("order_amt",  "요기요 주문금액",      ["주문금액", "주문 금액", "결제금액", "상품금액", "음식금액"]),
]
# 개념 → 손익 부호/VAT 기준
REVENUE = {"order_amt"}
COSTS = {"brokerage", "payment", "delivery", "ad", "discount"}
SKIP = {"deposit", "vat"}   # 입금액=검산용, 부가세=별도(P&L 제외)
BASIS = {  # vat_basis (실파일 확인 후 조정 가능)
    "order_amt": "gross", "brokerage": "gross", "payment": "gross",
    "delivery": "gross", "ad": "gross", "discount": "none",
}


def _norm(s) -> str:
    return str(s).replace(" ", "").strip() if s is not None else ""


class YogiyoImporter(DeliveryImporter):
    platform = "yogiyo"
    doc_type = "settlement"
    VERIFIED = False  # 실파일로 검증되면 True 로

    def _find_header(self, ws, scan=15):
        """요기요 키워드가 가장 많이 잡히는 행을 헤더로 추정. returns (row_idx0, {concept:col})."""
        all_syn = [(c, syn) for c, _, syns in SYNONYMS for syn in syns]
        best = (-1, -1, {})
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=scan, values_only=True)):
            colmap, claimed = {}, set()
            cells = [_norm(v) for v in row]
            for ci, cell in enumerate(cells):
                if not cell or ci in claimed:
                    continue
                for concept, _, syns in SYNONYMS:
                    if concept in colmap:
                        continue
                    if any(_norm(syn) in cell for syn in syns):
                        colmap[concept] = ci
                        claimed.add(ci)
                        break
            if len(colmap) > best[0]:
                best = (len(colmap), i, colmap)
        return best[1], best[2]

    def identify(self, filepath: str) -> bool:
        try:
            ws = openpyxl.load_workbook(filepath, read_only=True, data_only=True).active
        except Exception:
            return False
        _, colmap = self._find_header(ws)
        # 요기요 특징: 중개이용료 + 입금액(or 주문금액) 동시 존재
        return "brokerage" in colmap and ({"deposit", "order_amt"} & set(colmap))

    def extract(self, filepath: str) -> ImportResult:
        ws = openpyxl.load_workbook(filepath, read_only=True, data_only=True).active
        hrow, colmap = self._find_header(ws)
        won = self.won

        rows_iter = list(ws.iter_rows(min_row=hrow + 2, values_only=True))

        # 비용 컬럼 부호 자동판별: 합>0 이면 양수저장 → −로 환산
        sign = {}
        for concept in COSTS:
            ci = colmap.get(concept)
            if ci is None:
                continue
            tot = sum(won(r[ci]) for r in rows_iter if ci < len(r))
            sign[concept] = -1 if tot > 0 else 1

        recs, payout, period = [], 0.0, ""
        for r in rows_iter:
            anchor = colmap.get("order_amt", colmap.get("deposit"))
            if anchor is None or anchor >= len(r) or won(r[anchor]) == 0:
                # 날짜라도 있으면 진행, 전부 비면 skip
                if all(won(r[c]) == 0 for c in colmap.values() if c < len(r)):
                    continue
            # 기간 추정(첫 열에 날짜가 흔함)
            if not period:
                for cell in r[:3]:
                    s = str(cell or "")
                    if len(s) >= 7 and s[:4].isdigit():
                        period = s[:7]
                        break

            def emit(concept):
                ci = colmap.get(concept)
                if ci is None or ci >= len(r):
                    return
                val = won(r[ci])
                if not val:
                    return
                if concept in REVENUE:
                    amt = +val
                elif concept in COSTS:
                    amt = val * sign.get(concept, -1)
                else:
                    return
                name = next(n for c, n, _ in SYNONYMS if c == concept)
                recs.append(dict(date=period and f"{period}-01" or "", platform=self.platform,
                                 doc_type=self.doc_type, order_no="", item_name=name,
                                 amount=amt, vat_basis=BASIS.get(concept, "gross"), vat=0.0))

            for concept in (*REVENUE, *COSTS):
                emit(concept)
            dci = colmap.get("deposit")
            if dci is not None and dci < len(r):
                payout += won(r[dci])

        rows = pd.DataFrame(recs)

        # 자동 검산: 모든 항목 합 ≈ 입금액?
        note = "⚠️ 요기요 importer는 실파일 미검증(헤더 동의어 매칭). "
        note += f"인식 컬럼: {sorted(colmap.keys())}. "
        if payout and not rows.empty:
            net = rows["amount"].sum()
            diff = net - payout
            note += (f"검산: 항목합 {net:,.0f} vs 입금액 {payout:,.0f} "
                     f"(차이 {diff:,.0f}) → {'통과 ✅' if abs(diff) <= max(10, abs(payout)*0.02) else '불일치 ⚠️ SYNONYMS/부호 조정 필요'}.")
        else:
            note += "입금액 컬럼 미인식 → 검산 불가."

        return ImportResult(platform=self.platform, doc_type=self.doc_type,
                            period=period, rows=rows, payout_reported=payout, notes=note)
