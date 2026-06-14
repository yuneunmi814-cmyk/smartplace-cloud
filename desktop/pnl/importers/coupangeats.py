"""쿠팡이츠 정산 엑셀 importer — 2026-05 실파일로 검증 완료.

확정 사실 (손익계산서-agent-reference.md §3-F):
  - 진짜 OOXML xlsx, 단일시트, 3단 병합헤더(행0~2), 데이터 행3부터, 49컬럼.
  - 정산금액(45) = 주문금액(10) − 서비스이용료총액(33) − 최종광고비(42)   [행별·월합계 오차 0]
  - 서비스이용료총액(33) ≈ 중개이용료(16)+결제대행수수료(17+18)+배달비(21)  (VAT 조정뿐)
  - => 손익엔 개별수수료(16/17+18/21)+광고(42)로 분해. 33은 검증용으로만 (이중계상 금지).
  - txn_type(8): '결제'/'취소'. 취소행은 음수라 그냥 합산하면 상계됨.
  - 광고 그룹헤더 원본 오타: "촤종 광고비".
"""
from __future__ import annotations

import openpyxl
import pandas as pd

from .base import DeliveryImporter, ImportResult

# 0-based 컬럼 인덱스 (실파일 검증)
COL = dict(
    date=0, time=1, order_no=2, type=3, detail=4, store=6, pay_method=7,
    txn_type=8,
    order_amt=10,          # 주문금액 = 매출액 기준
    coupon_store=13,       # 쿠폰 상점부담 → 판매촉진비
    mid_fee=16,            # 중개이용료 산정후
    pg_base=17, pg_promo=18,   # 결제대행사 수수료 = 17+18
    deliv_fee=21,          # 배달비 산정후
    svc_total=33,          # 서비스이용료 산정후 총액(롤업) — 검증 전용
    svc_vat=32,            # 서비스이용료 산정후 부가세액 (번들 — 검증용)
    ad_supply=40, ad_vat=41, ad_final=42,   # 촤종 광고비: 공급가액/부가세/총액
    payout=45,             # 정산금액 산정후 (실입금, 검증용)
)
HEADER_GROUP_SIGNATURE = ["주문정보", "매출액", "중개이용료", "서비스이용료"]
DATA_START_ROW = 4  # openpyxl 1-based (헤더 3줄)


class CoupangEatsImporter(DeliveryImporter):
    platform = "coupangeats"
    doc_type = "settlement"

    def _load(self, filepath: str):
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        return wb.active

    def identify(self, filepath: str) -> bool:
        try:
            ws = self._load(filepath)
        except Exception:
            return False
        row0 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        # 병합셀 → forward-fill 후 시그니처 포함 검사
        seen, last = set(), ""
        for v in row0:
            if v is not None and str(v).strip():
                last = str(v).strip()
            seen.add(last)
        return all(sig in seen for sig in HEADER_GROUP_SIGNATURE)

    def extract(self, filepath: str) -> ImportResult:
        ws = self._load(filepath)
        won = self.won
        recs = []
        payout_total = 0.0
        period = None
        for r in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            if r[COL["order_no"]] in (None, ""):
                continue
            date = str(r[COL["date"]] or "")[:10]
            if period is None and len(date) >= 7:
                period = date[:7]
            order_no = str(r[COL["order_no"]])

            def emit(item, amt, basis="gross", vat=0.0):
                if amt:
                    recs.append(dict(date=date, platform=self.platform,
                                     doc_type=self.doc_type, order_no=order_no,
                                     item_name=item, amount=amt,
                                     vat_basis=basis, vat=vat))

            # 비용은 -won(): 결제행은 양수→음수(비용), 취소행은 음수→양수(환급 상계).
            # abs()를 쓰면 취소 환급이 상계되지 않아 비용이 부풀려진다.
            # 쿠팡 수수료는 총액(VAT포함) → gross. 광고는 파일에 부가세 명시 → exact.
            emit("쿠팡이츠 주문금액", +won(r[COL["order_amt"]]), "gross")
            emit("쿠팡이츠 중개이용료", -won(r[COL["mid_fee"]]), "gross")
            emit("쿠팡이츠 결제대행수수료",
                 -(won(r[COL["pg_base"]]) + won(r[COL["pg_promo"]])), "gross")
            emit("쿠팡이츠 배달비", -won(r[COL["deliv_fee"]]), "gross")
            emit("쿠팡이츠 광고비", -won(r[COL["ad_final"]]), "exact",
                 vat=-won(r[COL["ad_vat"]]))           # 광고 부가세 파일값(col41)
            emit("쿠팡이츠 쿠폰상점부담", -won(r[COL["coupon_store"]]), "none")
            payout_total += won(r[COL["payout"]])

        rows = pd.DataFrame(recs)
        return ImportResult(platform=self.platform, doc_type=self.doc_type,
                            period=period or "", rows=rows,
                            payout_reported=payout_total,
                            notes="서비스이용료(33)는 롤업이라 미계상; 개별수수료로 분해.")
