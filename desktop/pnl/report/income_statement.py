"""분류된 거래 → 손익계산서(당기순이익까지) 구조 + 엑셀 렌더."""
from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO

import pandas as pd


@dataclass
class IncomeStatement:
    period: str
    revenue: float
    cogs: float
    gross_profit: float
    sga_total: float
    sga_detail: dict          # {계정과목(l2): 금액}
    operating_profit: float
    nonop_income: float
    nonop_expense: float
    pretax_profit: float
    income_tax: float
    net_income: float
    platform_revenue: dict = field(default_factory=dict)  # 플랫폼별 매출
    warnings: list = field(default_factory=list)


def _sum(df: pd.DataFrame, l1: str) -> float:
    return float(df.loc[df["account_l1"] == l1, "amount"].sum())


def build_income_statement(df: pd.DataFrame, period: str = "") -> IncomeStatement:
    if df.empty:
        return IncomeStatement(period, 0, 0, 0, 0, {}, 0, 0, 0, 0, 0, 0)

    revenue = _sum(df, "매출액")                 # +
    cogs = -_sum(df, "매출원가")                 # 비용(음수) → 양수 표기
    gross = revenue - cogs

    sga = df[df["account_l1"] == "판매비와관리비"]
    sga_detail = (-sga.groupby("account_l2")["amount"].sum()).round(0).to_dict()
    sga_total = float(sum(sga_detail.values()))
    operating = gross - sga_total

    nonop_income = _sum(df, "영업외수익")
    nonop_expense = -_sum(df, "영업외비용")
    pretax = operating + nonop_income - nonop_expense
    tax = -_sum(df, "법인세등")
    net = pretax - tax

    platform_rev = (df[df["account_l1"] == "매출액"]
                    .groupby("platform")["amount"].sum().round(0).to_dict())

    return IncomeStatement(
        period=period, revenue=revenue, cogs=cogs, gross_profit=gross,
        sga_total=sga_total, sga_detail=sga_detail, operating_profit=operating,
        nonop_income=nonop_income, nonop_expense=nonop_expense,
        pretax_profit=pretax, income_tax=tax, net_income=net,
        platform_revenue=platform_rev,
    )


def as_lines(s: IncomeStatement) -> list[tuple[str, float, int]]:
    """(과목, 금액, 들여쓰기레벨) 순서대로. 표시·엑셀 공용."""
    lines = [
        ("Ⅰ. 매출액", s.revenue, 0),
    ]
    for plat, amt in s.platform_revenue.items():
        lines.append((f"    · {plat}", amt, 1))
    lines += [
        ("Ⅱ. 매출원가", s.cogs, 0),
        ("Ⅲ. 매출총이익", s.gross_profit, 0),
        ("Ⅳ. 판매비와관리비", s.sga_total, 0),
    ]
    for acc, amt in sorted(s.sga_detail.items(), key=lambda x: -x[1]):
        lines.append((f"    · {acc}", amt, 1))
    lines += [
        ("Ⅴ. 영업이익", s.operating_profit, 0),
        ("Ⅵ. 영업외수익", s.nonop_income, 0),
        ("Ⅶ. 영업외비용", s.nonop_expense, 0),
        ("Ⅷ. 법인세비용차감전순이익", s.pretax_profit, 0),
        ("Ⅸ. 법인세등", s.income_tax, 0),
        ("Ⅹ. 당기순이익", s.net_income, 0),
    ]
    return lines


def render_xlsx(s: IncomeStatement, vat_summary: dict | None = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "손익계산서"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18

    title = Font(bold=True, size=14)
    hdr = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="4472C4")
    bold = Font(bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(bottom=thin)
    right = Alignment(horizontal="right")

    ws["A1"] = f"손익계산서 ({s.period})"
    ws["A1"].font = title
    ws["A3"], ws["B3"] = "과목", "금액(원)"
    for c in ("A3", "B3"):
        ws[c].font = hdr; ws[c].fill = fill
    ws["B3"].alignment = right

    row = 4
    majors = {"Ⅰ", "Ⅱ", "Ⅲ", "Ⅳ", "Ⅴ", "Ⅵ", "Ⅶ", "Ⅷ", "Ⅸ", "Ⅹ"}
    totals = {"Ⅲ. 매출총이익", "Ⅴ. 영업이익", "Ⅷ. 법인세비용차감전순이익", "Ⅹ. 당기순이익"}
    for label, amount, level in as_lines(s):
        ws.cell(row, 1, label)
        cell = ws.cell(row, 2, round(amount))
        cell.number_format = "#,##0"
        cell.alignment = right
        if label in totals:
            ws.cell(row, 1).font = bold
            cell.font = bold
            ws.cell(row, 1).border = border
            cell.border = border
        elif label[:1] in majors:
            ws.cell(row, 1).font = Font(bold=True, color="333333")
        row += 1

    if vat_summary:
        row += 2
        ws.cell(row, 1, f"부가가치세 정산 ({vat_summary.get('label','')})").font = bold
        for k, lbl in (("output_vat", "매출세액"), ("input_vat", "매입세액(공제)"),
                       ("payable", "납부예상세액")):
            row += 1
            ws.cell(row, 1, lbl)
            c = ws.cell(row, 2, vat_summary.get(k, 0))
            c.number_format = "#,##0"; c.alignment = right
            if k == "payable":
                ws.cell(row, 1).font = bold; c.font = bold
        row += 1
        ws.cell(row, 1, f"· {vat_summary.get('note','')}").font = Font(size=9, color="888888")

    if s.warnings:
        row += 2
        ws.cell(row, 1, "비고/검증").font = bold
        for w in s.warnings:
            row += 1
            ws.cell(row, 1, f"· {w}")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
